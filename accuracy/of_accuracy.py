#!/usr/bin/env python3
import os
import glob
import argparse as ag
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

scoring = {
    'Center': [0, 0],
    'Left': [0, 1],
    'Right': [1, 0],
}

class Accuracy:
    def __init__(self, data):
        self.data = data
        self.point = ['Center', 'Left', 'Right']
        self.dir = ['In', 'Out']
        self.totals = {}
        self.scores = {}

        for p in self.point:
            for d in self.dir:
                e = p + d
                self.totals[e] = 0

        self.find_summary()

    def find_summary(self):
        # Locate "SUMMARY DATA" and extract the relevant section
        for idx, row in enumerate(self.data):
            if row[0] == 'SUMMARY DATA':
                summary_start = idx
                break
        else:
            raise ValueError("SUMMARY DATA not found in the file")

        summary_data = self.data[summary_start:]
        for idx, row in enumerate(summary_data):
            if row[0] == 'Condition':
                condition_start = idx
                break
        else:
            raise ValueError("Condition not found in the summary data")

        # Use only rows from the Condition onward
        self.data = summary_data[condition_start + 1:]
        self.headers = summary_data[condition_start]  # Set headers for reference

    def sum_stimulus(self, idx):
        event = self.data[idx][0]  # Get the 'Stimulus' column value
        presses = {}
        type = None
        direction = None

        for p in self.point:
            if p in event:
                type = p
                break
        for d in self.dir:
            if d in event:
                direction = d

        if type is None or direction is None:
            raise ValueError(f'Unrecognized event: {event}')

        event_key = type + direction
        n_events = int(self.data[idx][2])  # Get the 'n Events' column value
        n_c_event = 0
        i = 1

        while n_c_event != n_events:
            n_c_event = int(self.data[idx + i][2])
            presses[i] = n_c_event
            i += 1

        split = self.score_presses(presses, type, n_events)

        if event_key in self.scores:
            c_correct, c_incorrect = self.scores[event_key]
            self.scores[event_key] = (
                c_correct + split[0],
                c_incorrect + split[1]
            )
        else:
            self.scores[event_key] = split

        self.totals[event_key] = self.totals[event_key] + n_events

    def score_presses(self, presses, type, total):
        correct = 0
        incorrect = 0
        n_presses = 0

        for idx, press_count in presses.items():
            if idx == 3:
                continue
            weight = scoring[type]
            if idx - 1 < len(weight):
                if weight[idx - 1]:
                    correct += press_count
                else:
                    incorrect += press_count
            n_presses += press_count

        if type in ['Left', 'Right']:
            missed_presses = total - n_presses
            incorrect += missed_presses

        if type == 'Center':
            correct = total - incorrect

        return correct, incorrect

    def calculate_accuracy(self):
        for i, row in enumerate(self.data):
            if isinstance(row[0], str):
                self.sum_stimulus(i)
from openpyxl import Workbook

def process_directory(input_dir, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    for file_path in glob.glob(os.path.join(input_dir, '*.xlsx')):
        print(f"Processing {file_path}...")
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Extract data from the worksheet
        data = [list(row) for row in sheet.iter_rows(values_only=True)]

        acc = Accuracy(data)
        acc.calculate_accuracy()

        # Prepare output workbook
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Accuracy Scores"

        # Write headers
        output_sheet.append(['Event', 'Correct', 'Incorrect'])

        # Write accuracy scores
        for event, score in acc.scores.items():
            output_sheet.append([event, score[0], score[1]])

        # Save as Excel file
        output_file = os.path.join(output_dir, os.path.basename(file_path).replace('.xlsx', '_accuracy.xlsx'))
        output_workbook.save(output_file)
        print(f"Saved accuracy scores to {output_file}")

def main(args):
    process_directory(args.input_dir, args.output_dir)

if __name__ == '__main__':
    argv = ag.ArgumentParser()
    argv.add_argument('input_dir', type=str, help='Directory containing input Excel files')
    argv.add_argument('output_dir', type=str, help='Directory to save output CSV files')
    args = argv.parse_args()

    main(args)

